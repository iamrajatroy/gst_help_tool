"""
Exporter: generates Excel (.xlsx), CSV, and report files from GSTRecord datasets.

Uses openpyxl for formatted Excel output with summary and validation sheets.
Supports sheet chunking for very large datasets.
"""

from __future__ import annotations

import csv
import json
import logging
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


from gst_scraper.models import ExportConfig, ExportError, GSTRecord, ValidationReport

logger = logging.getLogger("gst_scraper.exporter")

# Excel styling constants
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_SUMMARY_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


class Exporter:
    """
    Exports GSTRecord datasets to Excel and CSV formats.

    Features:
    - Formatted Excel workbook with auto-width columns
    - Summary sheet with category distribution
    - Validation report sheet
    - CSV snapshot for machine processing
    - Sheet chunking for datasets exceeding Excel row limits
    """

    def __init__(self, config: ExportConfig | None = None):
        self.config = config or ExportConfig()
        self.output_dir = Path(self.config.output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_all(
        self,
        records: list[GSTRecord],
        validation_report: ValidationReport | None = None,
        run_summary: dict | None = None,
    ) -> dict[str, str]:
        """
        Export dataset to all configured formats.

        Returns:
            Dict of format → file path
        """
        outputs = {}

        # Excel
        try:
            xlsx_path = self.export_excel(records, validation_report, run_summary)
            outputs["xlsx"] = str(xlsx_path)
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            raise ExportError(self.config.excel_filename, str(e))

        # CSV
        try:
            csv_path = self.export_csv(records)
            outputs["csv"] = str(csv_path)
        except Exception as e:
            logger.error(f"CSV export failed: {e}")
            raise ExportError(self.config.csv_filename, str(e))

        # Validation report
        if validation_report:
            try:
                report_path = self.export_validation_report(validation_report)
                outputs["validation_report"] = str(report_path)
            except Exception as e:
                logger.warning(f"Validation report export failed: {e}")

        logger.info(
            f"Export complete: {', '.join(f'{k}={v}' for k, v in outputs.items())}",
            extra={"stage": "export"},
        )
        return outputs

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------

    def export_excel(
        self,
        records: list[GSTRecord],
        validation_report: ValidationReport | None = None,
        run_summary: dict | None = None,
    ) -> Path:
        """Export to formatted Excel workbook."""
        filepath = self.output_dir / self.config.excel_filename
        df = self._records_to_dataframe(records)

        # Check if we need chunked sheets
        if len(df) > self.config.chunk_size:
            return self._export_excel_chunked(df, filepath, validation_report, run_summary)

        wb = Workbook()

        # Main data sheet
        ws = wb.active
        ws.title = "GST Products"
        self._write_dataframe_to_sheet(ws, df)

        # Summary sheet
        if self.config.include_summary_sheet:
            ws_summary = wb.create_sheet("Summary")
            self._write_summary_sheet(ws_summary, df, run_summary)

        # Validation sheet
        if self.config.include_validation_sheet and validation_report:
            ws_val = wb.create_sheet("Validation Report")
            self._write_validation_sheet(ws_val, validation_report)

        wb.save(filepath)
        logger.info(f"Excel exported: {filepath} ({len(df)} rows)")
        return filepath

    def _export_excel_chunked(
        self,
        df: pd.DataFrame,
        filepath: Path,
        validation_report: ValidationReport | None,
        run_summary: dict | None,
    ) -> Path:
        """Export large datasets across multiple sheets."""
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        chunk_size = self.config.chunk_size
        num_chunks = (len(df) + chunk_size - 1) // chunk_size

        for i in range(num_chunks):
            start = i * chunk_size
            end = min(start + chunk_size, len(df))
            chunk = df.iloc[start:end]
            ws = wb.create_sheet(f"Data Part {i + 1}")
            self._write_dataframe_to_sheet(ws, chunk)

        # Summary
        if self.config.include_summary_sheet:
            ws_summary = wb.create_sheet("Summary")
            self._write_summary_sheet(ws_summary, df, run_summary)

        # Validation
        if self.config.include_validation_sheet and validation_report:
            ws_val = wb.create_sheet("Validation Report")
            self._write_validation_sheet(ws_val, validation_report)

        wb.save(filepath)
        logger.info(f"Chunked Excel exported: {filepath} ({len(df)} rows, {num_chunks} sheets)")
        return filepath

    def _write_dataframe_to_sheet(self, ws, df: pd.DataFrame) -> None:
        """Write DataFrame to worksheet with formatting."""
        # Headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(col_name))
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Data rows
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # Convert Decimal to float for Excel
                if isinstance(value, Decimal):
                    cell.value = float(value)
                    cell.number_format = '0.00'
                elif pd.isna(value):
                    cell.value = ""
                else:
                    cell.value = value
                cell.border = _BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=False)

        # Auto-width columns (sample-based for performance)
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = len(str(col_name))
            # Sample up to 100 rows for width calculation
            for row_idx in range(2, min(102, len(df) + 2)):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    max_len = max(max_len, min(len(str(cell_val)), 50))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 2

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-filter
        if len(df) > 0:
            ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(df.columns)).column_letter}{len(df) + 1}"

    def _write_summary_sheet(self, ws, df: pd.DataFrame, run_summary: dict | None = None) -> None:
        """Write summary statistics sheet."""
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

        row = 1

        # Title
        cell = ws.cell(row=row, column=1, value="GST Product Dataset — Summary")
        cell.font = Font(name="Calibri", size=14, bold=True)
        row += 2

        # General stats
        stats = [
            ("Generated At", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")),
            ("Total Records", len(df)),
            ("Unique HSN Codes", df["hsn_code"].nunique() if "hsn_code" in df.columns else "N/A"),
            ("", ""),
            ("--- Category Distribution ---", ""),
        ]

        if "top_category" in df.columns:
            for cat, count in df["top_category"].value_counts().items():
                stats.append((f"  {cat}", count))
            stats.append(("", ""))

        if "sub_category" in df.columns:
            stats.append(("--- Sub-Category Distribution ---", ""))
            for cat, count in df["sub_category"].value_counts().items():
                stats.append((f"  {cat}", count))
            stats.append(("", ""))

        if "confidence_flag" in df.columns:
            stats.append(("--- Confidence Distribution ---", ""))
            for flag, count in df["confidence_flag"].value_counts().items():
                stats.append((f"  {flag}", count))

        # Write stats
        for label, value in stats:
            ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=10, bold=bool(label.startswith("---")))
            ws.cell(row=row, column=2, value=value)
            row += 1

        # Run summary if provided
        if run_summary:
            row += 1
            ws.cell(row=row, column=1, value="--- Run Summary ---").font = Font(bold=True)
            row += 1
            for key, val in run_summary.items():
                if not isinstance(val, (dict, list)):
                    ws.cell(row=row, column=1, value=str(key))
                    ws.cell(row=row, column=2, value=str(val))
                    row += 1

    def _write_validation_sheet(self, ws, report: ValidationReport) -> None:
        """Write validation report sheet."""
        headers = ["Issue Type", "Severity", "Record ID", "HSN Code", "Description", "Source"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.border = _BORDER

        for row_idx, issue in enumerate(report.issues[:10000], 2):  # Cap at 10K issues
            ws.cell(row=row_idx, column=1, value=issue.issue_type).border = _BORDER
            ws.cell(row=row_idx, column=2, value=issue.severity).border = _BORDER
            ws.cell(row=row_idx, column=3, value=issue.record_id or "").border = _BORDER
            ws.cell(row=row_idx, column=4, value=issue.hsn_code or "").border = _BORDER
            ws.cell(row=row_idx, column=5, value=issue.description).border = _BORDER
            ws.cell(row=row_idx, column=6, value=issue.source_name or "").border = _BORDER

        # Auto-width
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 60
        ws.column_dimensions["F"].width = 25
        ws.freeze_panes = "A2"

    # ------------------------------------------------------------------
    # CSV export
    # ------------------------------------------------------------------

    def export_csv(self, records: list[GSTRecord]) -> Path:
        """Export to CSV."""
        filepath = self.output_dir / self.config.csv_filename
        df = self._records_to_dataframe(records)
        df.to_csv(filepath, index=False, encoding="utf-8-sig")
        logger.info(f"CSV exported: {filepath} ({len(df)} rows)")
        return filepath

    # ------------------------------------------------------------------
    # Validation report export
    # ------------------------------------------------------------------

    def export_validation_report(self, report: ValidationReport) -> Path:
        """Export validation report as CSV."""
        filepath = self.output_dir / "validation_report.csv"
        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=[
                "issue_type", "severity", "record_id", "hsn_code", "description", "source_name"
            ])
            writer.writeheader()
            for issue in report.issues:
                writer.writerow({
                    "issue_type": issue.issue_type,
                    "severity": issue.severity,
                    "record_id": issue.record_id or "",
                    "hsn_code": issue.hsn_code or "",
                    "description": issue.description,
                    "source_name": issue.source_name or "",
                })

        logger.info(f"Validation report exported: {filepath} ({len(report.issues)} issues)")
        return filepath

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _records_to_dataframe(records: list[GSTRecord]) -> pd.DataFrame:
        """Convert GSTRecord list to pandas DataFrame."""
        data = []
        for rec in records:
            data.append({
                "product_id": rec.product_id,
                "product_name": rec.product_name,
                "top_category": rec.top_category if isinstance(rec.top_category, str) else rec.top_category.value,
                "sub_category": rec.sub_category if isinstance(rec.sub_category, str) else rec.sub_category.value,
                "hsn_code": rec.hsn_code,
                "gst_description": rec.gst_description,
                "schedule": rec.schedule,
                "serial_no": rec.serial_no,
                "cgst_rate": float(Decimal(str(rec.cgst_rate))),
                "sgst_rate": float(Decimal(str(rec.sgst_rate))),
                "igst_rate": float(Decimal(str(rec.igst_rate))),
                "cess_rate": float(Decimal(str(rec.cess_rate))),
                "effective_from": rec.effective_from,
                "source_name": rec.source_name,
                "source_url": rec.source_url,
                "raw_text": rec.raw_text,
                "confidence_flag": rec.confidence_flag if isinstance(rec.confidence_flag, str) else rec.confidence_flag.value,
            })
        return pd.DataFrame(data)
